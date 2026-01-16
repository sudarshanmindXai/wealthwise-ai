import csv
import pandas as pd
import random
from datetime import datetime, timedelta
from pathlib import Path

# Constants for generating realistic data
BUSINESS_ENTITIES = [
    "Razorpay", "Stripe India", "Upwork Escrow", "Fiverr Int",
    "Tech Corp Client", "Global Consultancy", "Client Payment"
]

PERSONAL_ENTITIES = [
    "Swiggy", "Zomato", "Uber", "Ola",
    "Netflix", "Spotify", "Amazon Retail", "Flipkart",
    "Starbucks", "PVR Cinemas", "Apollo Pharmacy"
]

PII_ENTITIES = [
    "Transfer to 9876543210", "UPI-vikram@okicici", "Refund to ABCDE1234F",
    "Payment from 9988776655", "Loan EMI - Policy 1234567890"
]

def generate_transactions(count=50):
    start_date = datetime(2025, 4, 1)
    transactions = []
    
    for _ in range(count):
        txn_date = start_date + timedelta(days=random.randint(0, 300))
        date_str = txn_date.strftime("%d-%m-%Y")
        
        # Decide type
        rand_val = random.random()
        if rand_val < 0.2: # Business Income
            desc = random.choice(BUSINESS_ENTITIES)
            credit = round(random.uniform(10000, 150000), 2)
            debit = 0
            txn_type = "credit"
        elif rand_val < 0.7: # Personal Expense
            desc = random.choice(PERSONAL_ENTITIES)
            credit = 0
            debit = round(random.uniform(100, 5000), 2)
            txn_type = "debit"
        else: # PII / Ambiguous
            desc = random.choice(PII_ENTITIES)
            if "Transfer" in desc or "EMI" in desc:
                credit = 0
                debit = round(random.uniform(5000, 20000), 2)
                txn_type = "debit"
            else:
                credit = round(random.uniform(500, 5000), 2)
                debit = 0
                txn_type = "credit"
                
        transactions.append({
            "Date": date_str,
            "Description": desc,
            "Debit": debit,
            "Credit": credit,
            "Balance": 0 # Placeholder
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: datetime.strptime(x["Date"], "%d-%m-%Y"))
    balance = 50000.0
    for t in transactions:
        balance = balance + t["Credit"] - t["Debit"]
        t["Balance"] = round(balance, 2)
        
    return transactions

def save_samples():
    output_dir = Path("wealthwise/backend/sample_docs")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    data = generate_transactions(150) # More transactions for realism
    
    # --- Generate HDFC-style XLS/XLSX ---
    # Structure:
    # Rows 0-20: Metadata (Address, Account details, etc.)
    # Row 21: Header
    # Rows 22+: Data
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    
    # Metadata
    ws["A1"] = "HDFC BANK Ltd."
    ws["A2"] = "Statement of accounts"
    ws["A4"] = "From : 01/04/2025"
    ws["A5"] = "To : 31/03/2026"
    ws["A6"] = "Account No : 50100123456789"
    ws["A7"] = "Account Name : VIKRAM RATHORE"
    ws["A8"] = "Address : Flat 402, Oakwood Residency, Indiranagar, Bangalore"
    ws["A9"] = "City : BANGALORE"
    ws["A10"] = "State : KARNATAKA"
    ws["A11"] = "Pin : 560038"
    
    # Header Row (approx row 22)
    headers = ["Date", "Narration", "Chq./Ref.No.", "Value Dt", "Withdrawal Amt.", "Deposit Amt.", "Closing Balance"]
    header_row = 22
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=header)
        
    # Data
    for i, txn in enumerate(data, start=1):
        row = header_row + i
        ws.cell(row=row, column=1, value=txn["Date"])
        ws.cell(row=row, column=2, value=txn["Description"])
        ws.cell(row=row, column=3, value=f"N{random.randint(100000,999999)}") # Ref No
        ws.cell(row=row, column=4, value=txn["Date"]) # Value Dt same as Date
        
        # Withdrawal / Deposit
        if txn["Debit"] > 0:
             ws.cell(row=row, column=5, value=txn["Debit"])
             ws.cell(row=row, column=6, value=0.0) # Explicit 0.0 often used
        else:
             ws.cell(row=row, column=5, value=0.0)
             ws.cell(row=row, column=6, value=txn["Credit"])
             
        ws.cell(row=row, column=7, value=txn["Balance"])
        
    xlsx_path = output_dir / "bank_statement_vikram.xlsx"
    wb.save(xlsx_path)
    print(f"Generated {xlsx_path} (HDFC Style)")

    # --- Generate ICICI-style CSV ---
    # ICICI often has headers like: "Tran Date", "Chq No", "Particulars", "Debit", "Credit", "Balance"
    # Or "S No.", "Value Date", "Transaction Date", "Cheque Number", "Transaction Remarks", "Withdrawal Amount (INR)", "Deposit Amount (INR)", "Balance (INR)"
    
    csv_path = output_dir / "bank_statement_vikram.csv"
    
    # Headers mimicking ICICI
    csv_headers = ["S No.", "Value Date", "Transaction Date", "Cheque Number", "Transaction Remarks", "Withdrawal Amount (INR)", "Deposit Amount (INR)", "Balance (INR)"]
    
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        
        # ICICI sometimes has a few header lines too, but let's keep CSV standard for now 
        # or add a few lines to test "skiprows" robustness on CSV too.
        writer.writerow(["Op. Balance", "", "", "", "", "", "", "50000.00"])
        writer.writerow([]) # Blank line
        writer.writerow(csv_headers)
        
        for i, txn in enumerate(data, start=1):
            row = [
                i,
                txn["Date"], # Value Date
                txn["Date"], # Txn Date
                "-",         # Cheque No
                txn["Description"], # Remarks
                f"{txn['Debit']:.2f}" if txn['Debit'] > 0 else "0.00",
                f"{txn['Credit']:.2f}" if txn['Credit'] > 0 else "0.00",
                f"{txn['Balance']:.2f}"
            ]
            writer.writerow(row)
            
    print(f"Generated {csv_path} (ICICI Style)")

if __name__ == "__main__":
    from openpyxl import Workbook # Lazy import
    save_samples()
