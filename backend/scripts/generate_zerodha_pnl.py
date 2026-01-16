from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import random
from datetime import datetime, timedelta

def generate_zerodha_pnl(filename):
    wb = Workbook()
    
    # Defaults
    font_bold = Font(bold=True)
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # --- Sheet 1: Equity ---
    ws = wb.active
    ws.title = "Equity"
    
    # Metadata
    ws["B7"] = "Client ID"
    ws["C7"] = "VKR123"
    ws["B8"] = "Client Name"
    ws["C8"] = "VIKRAM RATHORE"
    ws["B9"] = "PAN"
    ws["C9"] = "ABCDE1234F"
    
    ws["B12"] = "Realized Profit Breakdown"
    ws["B12"].font = font_bold
    
    # Header Row 14 (1-based index 15) -> Python Row 15
    # The reference said Row 14 in pandas (0-indexed) => Row 15 in Excel (1-indexed)
    header_row = 15
    headers = ["Symbol", "ISIN", "Quantity", "Buy Value", "Sell Value", "Realized Profit", "Realized Profit %"]
    
    for i, h in enumerate(headers, start=2): # Start Col B (2)
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = font_bold
        c.fill = header_fill

    # Data
    scripts = ["RELIANCE", "TCS", "INFY", "HDFCBANK", "ICICIBANK", "SBIN"]
    row_idx = header_row + 1
    
    for scrip in scripts:
        qty = random.randint(10, 100)
        buy_price = random.uniform(500, 2500)
        sell_price = buy_price * random.uniform(0.9, 1.3)
        
        buy_val = qty * buy_price
        sell_val = qty * sell_price
        profit = sell_val - buy_val
        pct = (profit / buy_val) * 100
        
        ws.cell(row=row_idx, column=2, value=scrip)
        ws.cell(row=row_idx, column=3, value=f"INE{random.randint(100,999)}A010{random.randint(10,99)}")
        ws.cell(row=row_idx, column=4, value=qty)
        ws.cell(row=row_idx, column=5, value=round(buy_val, 2))
        ws.cell(row=row_idx, column=6, value=round(sell_val, 2))
        ws.cell(row=row_idx, column=7, value=round(profit, 2))
        ws.cell(row=row_idx, column=8, value=round(pct, 2))
        
        row_idx += 1
        
    # --- Sheet 2: Equity Dividends ---
    ws_div = wb.create_sheet("Equity Dividends")
    
    # Metadata
    ws_div["B7"] = "Client ID"; ws_div["C7"] = "VKR123"
    ws_div["B8"] = "Client Name"; ws_div["C8"] = "VIKRAM RATHORE"
    ws_div["B9"] = "PAN"; ws_div["C9"] = "ABCDE1234F"
    
    ws_div["B10"] = "Equity Dividends from 01/04/2025 to 31/03/2026"
    
    # Header Row 15
    headers_div = ["Symbol", "ISIN", "Date", "Quantity", "Net Dividend Amount"]
    for i, h in enumerate(headers_div, start=2):
        c = ws_div.cell(row=header_row, column=i, value=h)
        c.font = font_bold
        c.fill = header_fill
        
    # Data
    row_idx = header_row + 1
    for scrip in ["ITC", "VEDL", "IOC"]:
        qty = random.randint(50, 200)
        div = random.uniform(5, 20)
        amt = qty * div
        
        ws_div.cell(row=row_idx, column=2, value=scrip)
        ws_div.cell(row=row_idx, column=3, value=f"INE{random.randint(100,999)}A010{random.randint(10,99)}")
        ws_div.cell(row=row_idx, column=4, value="2025-08-15")
        ws_div.cell(row=row_idx, column=5, value=qty)
        ws_div.cell(row=row_idx, column=6, value=round(amt, 2))
        row_idx += 1

    wb.save(filename)
    print(f"Generated {filename}")

if __name__ == "__main__":
    output_dir = Path("wealthwise/backend/sample_docs")
    output_dir.mkdir(parents=True, exist_ok=True)
    generate_zerodha_pnl(str(output_dir / "Zerodha_pnl_vikram.xlsx"))
